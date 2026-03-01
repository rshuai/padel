from __future__ import annotations

import csv
import shutil
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..config import settings
from ..models import (
    ArtifactType,
    ControlException,
    Entity,
    FXRate,
    OutputArtifact,
    UploadedFile,
)


def _artifact_path(engagement_id: str, filename: str) -> Path:
    folder = settings.outputs_dir / engagement_id
    folder.mkdir(parents=True, exist_ok=True)
    return folder / filename


def reset_output_artifacts(db: Session, engagement_id: str) -> None:
    folder = settings.outputs_dir / engagement_id
    if folder.exists():
        shutil.rmtree(folder)
    folder.mkdir(parents=True, exist_ok=True)

    db.execute(delete(OutputArtifact).where(OutputArtifact.engagement_id == engagement_id))


def _register_artifact(db: Session, engagement_id: str, artifact_type: str, path: Path) -> None:
    db.add(
        OutputArtifact(
            engagement_id=engagement_id,
            artifact_type=artifact_type,
            file_path=str(path),
        )
    )


def _write_csv(path: Path, header: list[str], rows: list[list]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def write_consolidated_tb(db: Session, engagement_id: str, consolidated_rows: list[dict]) -> Path:
    path = _artifact_path(engagement_id, "consolidated_trial_balance_usd.csv")
    rows = [
        [
            row["group_account_code"],
            row.get("group_account_name") or "",
            row.get("account_class") or "",
            str(Decimal(str(row["usd_amount"])).quantize(Decimal("0.01"))),
        ]
        for row in consolidated_rows
    ]
    _write_csv(path, ["Group Account Code", "Group Account Name", "Account Class", "USD Amount"], rows)
    _register_artifact(db, engagement_id, ArtifactType.CONSOLIDATED_TB.value, path)
    return path


def write_income_statement(db: Session, engagement_id: str, consolidated_rows: list[dict]) -> Path:
    path = _artifact_path(engagement_id, "consolidated_income_statement_usd.csv")
    rows = [
        [
            row["group_account_code"],
            row.get("group_account_name") or "",
            str(Decimal(str(row["usd_amount"])).quantize(Decimal("0.01"))),
        ]
        for row in consolidated_rows
        if row.get("account_class") in {"REVENUE", "EXPENSE"}
    ]
    _write_csv(path, ["Group Account Code", "Group Account Name", "USD Amount"], rows)
    _register_artifact(db, engagement_id, ArtifactType.INCOME_STATEMENT.value, path)
    return path


def write_balance_sheet(db: Session, engagement_id: str, consolidated_rows: list[dict]) -> Path:
    path = _artifact_path(engagement_id, "consolidated_balance_sheet_usd.csv")
    rows = [
        [
            row["group_account_code"],
            row.get("group_account_name") or "",
            str(Decimal(str(row["usd_amount"])).quantize(Decimal("0.01"))),
        ]
        for row in consolidated_rows
        if row.get("account_class") in {"ASSET", "LIABILITY", "EQUITY"}
    ]
    _write_csv(path, ["Group Account Code", "Group Account Name", "USD Amount"], rows)
    _register_artifact(db, engagement_id, ArtifactType.BALANCE_SHEET.value, path)
    return path


def write_fx_report(db: Session, engagement_id: str) -> Path:
    path = _artifact_path(engagement_id, "fx_translation_report.csv")

    entities = {
        entity.id: entity.name
        for entity in db.execute(select(Entity).where(Entity.engagement_id == engagement_id)).scalars().all()
    }
    fx_rows = db.execute(
        select(FXRate).where(FXRate.engagement_id == engagement_id).order_by(FXRate.entity_id, FXRate.rate_type, FXRate.rate_date)
    ).scalars().all()
    rows = [
        [
            entities.get(row.entity_id, row.entity_id),
            row.base_currency,
            row.quote_currency,
            row.rate_type,
            row.rate_date.isoformat(),
            str(Decimal(row.rate).quantize(Decimal("0.00000001"))),
            row.source,
            row.methodology,
            row.missing_days,
            "Y" if row.is_override else "N",
            row.note or "",
        ]
        for row in fx_rows
    ]

    _write_csv(
        path,
        [
            "Entity",
            "From Currency",
            "To Currency",
            "Rate Type",
            "Rate Date",
            "Rate",
            "Source",
            "Methodology",
            "Missing Days",
            "Manual Override",
            "Notes",
        ],
        rows,
    )
    _register_artifact(db, engagement_id, ArtifactType.FX_REPORT.value, path)
    return path


def write_exception_report(db: Session, engagement_id: str) -> Path:
    path = _artifact_path(engagement_id, "exception_validation_report.csv")
    exceptions = db.execute(
        select(ControlException)
        .where(ControlException.engagement_id == engagement_id)
        .order_by(ControlException.blocking.desc(), ControlException.created_at)
    ).scalars().all()
    rows = [
        [
            row.severity,
            "BLOCKING" if row.blocking else "NON-BLOCKING",
            row.category,
            row.message,
            row.entity_id or "",
            row.account_code or "",
            row.status,
        ]
        for row in exceptions
    ]
    _write_csv(
        path,
        ["Severity", "Blocking", "Category", "Message", "Entity ID", "Account Code", "Status"],
        rows,
    )
    _register_artifact(db, engagement_id, ArtifactType.EXCEPTION_REPORT.value, path)
    return path


def write_journal_report(db: Session, engagement_id: str) -> Path:
    from ..models import ConsolidationJournal

    path = _artifact_path(engagement_id, "consolidation_journal_log.csv")
    journals = db.execute(
        select(ConsolidationJournal)
        .where(ConsolidationJournal.engagement_id == engagement_id)
        .order_by(ConsolidationJournal.created_at, ConsolidationJournal.id)
    ).scalars().all()

    rows = [
        [
            row.created_at.isoformat(),
            row.journal_type,
            row.description,
            row.debit_account,
            row.credit_account,
            str(Decimal(row.amount_usd).quantize(Decimal("0.01"))),
            row.entity_id or "",
            row.source_reference or "",
            row.created_by,
        ]
        for row in journals
    ]
    _write_csv(
        path,
        [
            "Timestamp",
            "Journal Type",
            "Description",
            "Debit Account",
            "Credit Account",
            "Amount USD",
            "Entity ID",
            "Source",
            "Created By",
        ],
        rows,
    )
    _register_artifact(db, engagement_id, ArtifactType.JOURNAL_LOG.value, path)
    return path


def _normalize_header(cell_value: object) -> str:
    return "".join(ch for ch in str(cell_value).strip().lower() if ch.isalnum())


def _populate_template_with_values(template_path: Path, output_path: Path, consolidated_map: dict[str, Decimal]) -> None:
    wb = load_workbook(template_path)

    for ws in wb.worksheets:
        header_row = None
        code_col = None
        amount_col = None

        for row_idx in range(1, min(ws.max_row, 30) + 1):
            row_cells = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, min(ws.max_column, 20) + 1)]
            normalized = [_normalize_header(value) for value in row_cells]

            for idx, val in enumerate(normalized, start=1):
                if val in {"groupaccountcode", "accountcode", "code"}:
                    code_col = idx
                if val in {"usdamount", "amountusd", "consolidatedusd", "amount"}:
                    amount_col = idx

            if code_col and amount_col:
                header_row = row_idx
                break

        if not (header_row and code_col and amount_col):
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            code = ws.cell(row=row_idx, column=code_col).value
            if code is None:
                continue
            code_text = str(code).strip()
            if not code_text:
                continue
            if code_text in consolidated_map:
                ws.cell(row=row_idx, column=amount_col).value = float(
                    consolidated_map[code_text].quantize(Decimal("0.01"))
                )

    wb.save(output_path)


def write_reporting_pack_from_template(
    db: Session,
    engagement_id: str,
    consolidated_rows: list[dict],
) -> Path | None:
    template = db.execute(
        select(UploadedFile)
        .where(
            UploadedFile.engagement_id == engagement_id,
            UploadedFile.file_type == "CONSOLIDATION_TEMPLATE",
        )
        .order_by(UploadedFile.uploaded_at.desc())
    ).scalar_one_or_none()

    if not template:
        return None

    template_path = Path(template.storage_path)
    if template_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return None

    output_path = _artifact_path(engagement_id, "consolidated_reporting_pack.xlsx")
    consolidated_map = {
        row["group_account_code"]: Decimal(str(row["usd_amount"]))
        for row in consolidated_rows
    }
    _populate_template_with_values(template_path, output_path, consolidated_map)
    _register_artifact(db, engagement_id, ArtifactType.REPORTING_PACK.value, output_path)
    return output_path


def generate_all_outputs(db: Session, engagement_id: str, consolidated_rows: list[dict]) -> list[Path]:
    reset_output_artifacts(db, engagement_id)

    files = [
        write_consolidated_tb(db, engagement_id, consolidated_rows),
        write_income_statement(db, engagement_id, consolidated_rows),
        write_balance_sheet(db, engagement_id, consolidated_rows),
        write_journal_report(db, engagement_id),
        write_fx_report(db, engagement_id),
        write_exception_report(db, engagement_id),
    ]

    pack = write_reporting_pack_from_template(db, engagement_id, consolidated_rows)
    if pack:
        files.append(pack)

    return files
